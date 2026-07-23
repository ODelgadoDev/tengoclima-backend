import csv
from io import BytesIO, StringIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .services import agrupar_movimientos, resumir_movimientos


AZUL = "17445A"
AZUL_CLARO = "DDEBF2"
NARANJA = "F5822A"
BLANCO = "FFFFFF"
GRIS = "E2E8F0"
VERDE = "166534"
ROJO = "B91C1C"
MONEDA = '$#,##0.00;[Red]-$#,##0.00'
FECHA = "dd/mm/yyyy"


MOVIMIENTOS_HEADERS = [
    "Fecha",
    "Tipo",
    "Concepto",
    "Categoría",
    "Cliente",
    "Proyecto",
    "Cotización",
    "Factura",
    "Método de pago",
    "Referencia",
    "Proveedor",
    "Subtotal",
    "IVA",
    "Monto",
    "Notas",
    "Registrado por",
]


def _fila_movimiento(movimiento):
    return [
        movimiento["fecha"],
        movimiento["tipo"],
        movimiento["concepto"],
        movimiento["categoria_nombre"],
        movimiento["cliente_nombre"],
        movimiento["proyecto_nombre"],
        movimiento["cotizacion_codigo"],
        movimiento["factura_folio"],
        movimiento["metodo_pago"],
        movimiento["referencia"],
        movimiento["proveedor"],
        movimiento["subtotal"],
        movimiento["iva"],
        movimiento["monto"],
        movimiento["notas"],
        movimiento["creado_por"],
    ]


def _estilizar_tabla(sheet, headers, currency_columns=(), date_columns=()):
    header_fill = PatternFill("solid", fgColor=AZUL)
    header_font = Font(color=BLANCO, bold=True)
    border = Border(bottom=Side(style="thin", color=GRIS))

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column in currency_columns:
        for cell in sheet[column][1:]:
            cell.number_format = MONEDA

    for column in date_columns:
        for cell in sheet[column][1:]:
            cell.number_format = FECHA

    for index, header in enumerate(headers, start=1):
        values = [str(header)]
        for row in sheet.iter_rows(
            min_row=2,
            min_col=index,
            max_col=index,
        ):
            value = row[0].value
            if value is not None:
                values.append(str(value))
        width = min(max(len(value) for value in values) + 2, 38)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 11)


def _crear_hoja_movimientos(workbook, title, movimientos):
    sheet = workbook.create_sheet(title)
    sheet.append(MOVIMIENTOS_HEADERS)
    for movimiento in movimientos:
        sheet.append(_fila_movimiento(movimiento))
    _estilizar_tabla(
        sheet,
        MOVIMIENTOS_HEADERS,
        currency_columns=("L", "M", "N"),
        date_columns=("A",),
    )
    return sheet


def _crear_hoja_resumen(workbook, movimientos, filtros):
    resumen = resumir_movimientos(movimientos)
    sheet = workbook.active
    sheet.title = "Resumen"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:D1")
    sheet["A1"] = "TENGOCLIMA — Libro contable"
    sheet["A1"].fill = PatternFill("solid", fgColor=AZUL)
    sheet["A1"].font = Font(color=BLANCO, bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 30

    sheet["A3"] = "Generado"
    sheet["B3"] = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    sheet["A4"] = "Periodo"
    desde = filtros.get("fecha_desde") or "Inicio"
    hasta = filtros.get("fecha_hasta") or "Actualidad"
    sheet["B4"] = f"{desde} a {hasta}"
    sheet["A5"] = "Tipo"
    sheet["B5"] = filtros.get("tipo") or "Todos"
    sheet["A6"] = "Búsqueda"
    sheet["B6"] = filtros.get("search") or "Sin búsqueda"

    metrics = [
        ("Ingresos", resumen["ingresos"], VERDE),
        ("Gastos", resumen["gastos"], ROJO),
        ("Utilidad", resumen["utilidad"], AZUL),
        ("IVA de ingresos", resumen["iva_ingresos"], NARANJA),
        ("IVA de gastos", resumen["iva_gastos"], NARANJA),
        ("IVA neto", resumen["iva_neto"], AZUL),
        ("Movimientos", resumen["movimientos"], AZUL),
    ]

    row = 9
    for label, value, color in metrics:
        sheet[f"A{row}"] = label
        sheet[f"A{row}"].font = Font(bold=True, color=AZUL)
        sheet[f"B{row}"] = value
        sheet[f"B{row}"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
        sheet[f"B{row}"].font = Font(bold=True, color=color)
        if label != "Movimientos":
            sheet[f"B{row}"].number_format = MONEDA
        row += 1

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 4
    sheet.column_dimensions["D"].width = 4


def _crear_hoja_agrupada(workbook, title, grupos, entity_label):
    sheet = workbook.create_sheet(title)
    headers = [
        entity_label,
        "Movimientos",
        "Ingresos",
        "Gastos",
        "Utilidad",
        "IVA ingresos",
        "IVA gastos",
        "IVA neto",
    ]
    sheet.append(headers)
    for grupo in grupos:
        sheet.append(
            [
                grupo["nombre"],
                grupo["movimientos"],
                grupo["ingresos"],
                grupo["gastos"],
                grupo["utilidad"],
                grupo["iva_ingresos"],
                grupo["iva_gastos"],
                grupo["iva_neto"],
            ],
        )
    _estilizar_tabla(
        sheet,
        headers,
        currency_columns=("C", "D", "E", "F", "G", "H"),
    )


def construir_excel(movimientos, filtros):
    workbook = Workbook()
    _crear_hoja_resumen(workbook, movimientos, filtros)
    _crear_hoja_movimientos(workbook, "Movimientos", movimientos)
    _crear_hoja_movimientos(
        workbook,
        "Ingresos",
        [m for m in movimientos if m["tipo"] == "INGRESO"],
    )
    _crear_hoja_movimientos(
        workbook,
        "Gastos",
        [m for m in movimientos if m["tipo"] == "GASTO"],
    )
    _crear_hoja_agrupada(
        workbook,
        "Por proyecto",
        agrupar_movimientos(
            movimientos,
            "proyecto_id",
            "proyecto_nombre",
        ),
        "Proyecto",
    )
    _crear_hoja_agrupada(
        workbook,
        "Por cliente",
        agrupar_movimientos(
            movimientos,
            "cliente_id",
            "cliente_nombre",
        ),
        "Cliente",
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def construir_csv(movimientos):
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(MOVIMIENTOS_HEADERS)
    for movimiento in movimientos:
        writer.writerow(_fila_movimiento(movimiento))
    return "\ufeff" + output.getvalue()
